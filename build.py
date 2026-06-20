# -*- coding: utf-8 -*-
"""
產生「領藥（調劑）查詢工具」單一 HTML。
特性：
  * HTML 內【不含】任何個人健康資料。
  * 開頁後由使用者點「匯入查詢來源」選擇 DRUGT.xml（健保調劑申報檔，Big5），
    於瀏覽器端即時解析，再以身分證字號查詢。
  * 藥名對照（藥品代碼→中文藥名）與院所對照（院所代碼→名稱）為公開參考資料，
    建置時嵌入頁面；找不到對照檔時仍可運作（只顯示代碼）。
輸入檔放在同資料夾：
  - 健保用藥*歷史資料*.xlsx  （藥名）
  - *院所名冊*.csv / *醫事機構*.csv|xlsx （院所名，可選）
"""
import json, os, glob

OUT = "領藥查詢.html"
NAME_XLSX_GLOB = "健保用藥*歷史資料*.xlsx"


def load_drug_names():
    files = sorted(glob.glob(NAME_XLSX_GLOB))
    if not files:
        print("（提示）找不到藥名 Excel，藥名欄將只顯示代碼。")
        return {}
    path = files[-1]
    try:
        import openpyxl
    except ImportError:
        print("（提示）未安裝 openpyxl，pip install openpyxl")
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows)
    def col(names):
        for i, h in enumerate(header):
            if h and str(h).strip() in names:
                return i
        return None
    ci = col({"藥品代碼", "代碼"}); ni = col({"中文藥名", "藥品名稱", "中文品名"})
    if ci is None or ni is None: ci, ni = 0, 1
    m = {}
    for r in rows:
        if ci < len(r) and ni < len(r) and r[ci]:
            m[str(r[ci]).strip().upper()] = (str(r[ni]).strip() if r[ni] else "")
    print(f"載入藥名對照：{os.path.basename(path)}（{len(m)} 筆）")
    return m


def load_hosp_names():
    cands = []
    for pat in ("*院所名冊*.csv", "*院所*.csv", "*醫事機構*.csv",
                "*院所名冊*.xlsx", "*醫事機構*.xlsx"):
        cands += glob.glob(pat)
    cands = sorted(set(cands))
    if not cands:
        print("（提示）找不到院所名冊，院所欄將只顯示代碼。可至 data.gov.tw/dataset/168341 下載。")
        return {}
    path = cands[-1]
    if path.lower().endswith(".csv"):
        import csv, io
        data = open(path, "rb").read()
        for enc in ("utf-8-sig", "big5", "utf-8"):
            try: text = data.decode(enc); break
            except UnicodeDecodeError: text = data.decode("big5", "replace")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
    if not rows: return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    def col(names):
        for i, h in enumerate(header):
            if h in names: return i
        return None
    ci = col({"HOSP_ID", "醫事機構代碼", "院所代碼", "代碼"})
    ni = col({"HOSP_NAME", "醫事機構名稱", "院所名稱", "名稱"})
    if ci is None or ni is None: ci, ni = 1, 3
    m = {}
    for r in rows[1:]:
        if ci < len(r) and ni < len(r) and r[ci]:
            code = str(r[ci]).strip().upper()
            name = str(r[ni]).strip() if r[ni] is not None else ""
            if code and name: m[code] = name
    print(f"載入院所名冊：{os.path.basename(path)}（{len(m)} 筆）")
    return m


def main():
    drug_names = load_drug_names()
    hosp_names = load_hosp_names()
    page = HTML_TEMPLATE \
        .replace("__DRUGS__", json.dumps(drug_names, ensure_ascii=False)) \
        .replace("__HOSPS__", json.dumps(hosp_names, ensure_ascii=False)) \
        .replace("__NDRUG__", str(len(drug_names))) \
        .replace("__NHOSP__", str(len(hosp_names)))
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(page)
    print(f"完成：{OUT}（藥名 {len(drug_names)} 筆 / 院所 {len(hosp_names)} 筆，不含個資）")


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>領藥（調劑）日期查詢</title>
<style>
  :root{ --bg:#0f172a; --card:#1e293b; --line:#334155; --txt:#e2e8f0;
         --muted:#94a3b8; --accent:#38bdf8; --good:#34d399; --warn:#fbbf24; }
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,"PingFang TC","Microsoft JhengHei",sans-serif;
       background:var(--bg);color:var(--txt);padding:24px;line-height:1.5}
  .wrap{max-width:880px;margin:0 auto}
  h1{font-size:20px;margin:0 0 4px}
  .sub{color:var(--muted);font-size:13px;margin-bottom:16px}
  .src{background:var(--card);border:1px dashed var(--line);border-radius:12px;
       padding:14px 16px;margin-bottom:18px}
  .src .row{display:flex;gap:12px;align-items:center;flex-wrap:wrap}
  .filebtn{padding:10px 18px;border:0;border-radius:10px;background:var(--accent);
           color:#06283d;font-weight:700;font-size:15px;cursor:pointer}
  .filebtn:hover{filter:brightness(1.08)}
  .status{font-size:13px;color:var(--muted)}
  .status.ok{color:var(--good)} .status.warn{color:var(--warn)}
  .clr{padding:6px 12px;border:1px solid var(--line);border-radius:8px;background:transparent;
       color:var(--muted);font-size:13px;cursor:pointer}
  .clr:hover{color:var(--txt);border-color:var(--accent)}
  .hint{margin-top:8px;color:var(--muted);font-size:12px}
  .bar{display:flex;gap:10px;flex-wrap:wrap;opacity:.5;pointer-events:none;transition:.2s}
  .bar.on{opacity:1;pointer-events:auto}
  input[type=text]{flex:1;min-width:220px;padding:12px 14px;border-radius:10px;border:1px solid var(--line);
        background:var(--card);color:var(--txt);font-size:16px;letter-spacing:1px}
  input[type=text]:focus{outline:none;border-color:var(--accent)}
  button.go{padding:12px 20px;border:0;border-radius:10px;background:var(--accent);
         color:#06283d;font-weight:700;font-size:15px;cursor:pointer}
  #out{margin-top:22px}
  .person{font-size:16px;margin:0 0 12px;color:var(--accent)}
  .person b{color:var(--txt)}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 16px;margin-bottom:12px}
  .date{font-size:18px;font-weight:700;color:var(--good)}
  .time{color:var(--warn);font-weight:800;font-size:18px;margin-left:4px}
  .meta{color:var(--muted);font-size:13px;margin-top:4px}
  .meta.rx{color:var(--accent);font-weight:700}
  table{width:100%;border-collapse:collapse;margin-top:10px;font-size:13px}
  th,td{text-align:left;padding:6px 8px;border-bottom:1px solid var(--line);vertical-align:top}
  th{color:var(--muted);font-weight:600}
  .dname{color:var(--txt);font-weight:600}
  .empty{color:var(--muted);padding:20px;text-align:center}
  .tag{display:inline-block;background:#0b1220;border:1px solid var(--line);border-radius:6px;padding:1px 7px;font-size:12px;color:var(--muted)}
  .note{margin-top:24px;color:var(--muted);font-size:12px;border-top:1px solid var(--line);padding-top:12px}
</style>
</head>
<body>
<div class="wrap">
  <h1>領藥（調劑）日期查詢</h1>
  <div class="sub">內建藥名對照 __NDRUG__ 筆 / 院所對照 __NHOSP__ 筆　•　本頁不含個資，需自行匯入申報檔　•　申報檔僅記錄到「日」；另可匯入每日領藥紀錄檔以補上精確時:分</div>

  <div class="src">
    <div class="row">
      <input type="file" id="file" accept=".xml,.zip" style="display:none">
      <button class="filebtn" onclick="document.getElementById('file').click()">📂 匯入查詢來源（DRUGT.xml / .zip）</button>
      <span class="status" id="st">尚未載入　•　請選擇健保調劑申報檔 DRUGT.xml 或 DRUGT.zip</span>
      <button class="clr" id="clrBtn" onclick="clearSrc()" style="display:none">清除預設</button>
    </div>
    <div class="hint">匯入後會記住此來源，下次開啟自動載入；選新檔即覆蓋。資料僅存於本機瀏覽器。</div>
  </div>

  <div class="src">
    <div class="row">
      <input type="file" id="tdir" webkitdirectory directory multiple style="display:none">
      <input type="file" id="tfile" accept=".txt,.zip" multiple style="display:none">
      <button class="filebtn" onclick="document.getElementById('tdir').click()">📁 匯入整個資料夾（如 sale）</button>
      <button class="clr" onclick="document.getElementById('tfile').click()">或選單一／多個 S 檔</button>
      <span class="status" id="tst">尚未載入領藥時間檔（選用）　•　含每筆精確領藥時:分，可累積多日</span>
      <button class="clr" id="tclrBtn" onclick="clearTimes()" style="display:none">清除時間檔</button>
    </div>
    <div class="hint">每日領藥紀錄檔（如 S1150618.txt，Big5）內含精確領藥時間。可整個資料夾匯入（自動挑出其中的 .txt／.zip、略過其他檔），或多次匯入累積；查詢時會與申報檔合併並自動去重。資料僅存於本機瀏覽器。</div>
  </div>

  <div class="bar" id="bar">
    <input type="text" id="q" placeholder="輸入身分證字號，例如 A123456789" autocomplete="off">
    <button class="go" onclick="run()">查詢</button>
  </div>

  <div id="out"></div>
  <div class="note">※ 申報檔含個人健康資料，僅於本機瀏覽器內解析，不會上傳。請勿外流原始檔。</div>
</div>
<script>
const DRUGS = __DRUGS__;     // 代碼 -> 中文藥名
const HOSPS = __HOSPS__;     // 院所代碼 -> 名稱
const DEPT = {"00":"家醫","01":"內科","02":"外科","03":"兒科","04":"婦產","05":"骨科",
  "06":"神經","07":"精神","08":"皮膚","09":"泌尿","10":"耳鼻喉","11":"眼科","12":"復健",
  "13":"麻醉","14":"放射","23":"急診"};
const esc = s => (s||"").replace(/[&<>]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;"}[c]));
let INDEX = null;
let STIMES_BY_ID = {};   // 身分證 -> [每日領藥紀錄(含時:分)]
let STIME_SEEN = {};     // 去重用 uid 集合

function rocToAd(s){
  s=(s||"").trim();
  if(!/^\d+$/.test(s) || s.length<5) return s;
  s=s.padStart(7,"0");
  return (parseInt(s.slice(0,-4),10)+1911)+"-"+s.slice(-4,-2)+"-"+s.slice(-2);
}
function tag(block,name){
  const m=block.match(new RegExp("<"+name+">([\\s\\S]*?)<\\/"+name+">"));
  return m?m[1].trim():"";
}
function parseXML(text){
  const idx={};
  const dds=text.match(/<ddata>[\s\S]*?<\/ddata>/g)||[];
  let nrec=0;
  for(const blk of dds){
    const drugs=[];
    const pds=blk.match(/<pdata>[\s\S]*?<\/pdata>/g)||[];
    for(const p of pds){
      const code=tag(p,"p2");
      drugs.push({kind:tag(p,"p1"),code:code,name:DRUGS[code.toUpperCase()]||"",
        dose:tag(p,"p3"),freq:tag(p,"p4"),route:tag(p,"p5"),
        days:tag(p,"p11"),qty:tag(p,"p7")});
    }
    const id=tag(blk,"d3").toUpperCase();
    const rec={id:id,name:tag(blk,"d20"),birth:rocToAd(tag(blk,"d6")),
      visit:rocToAd(tag(blk,"d14")),dispense:rocToAd(tag(blk,"d23")),
      dept:tag(blk,"d22"),hosp:tag(blk,"d21"),
      hospname:HOSPS[tag(blk,"d21").toUpperCase()]||"",
      points:tag(blk,"d18"),drugs:drugs};
    (idx[id]=idx[id]||[]).push(rec); nrec++;
  }
  for(const k in idx) idx[k].sort((a,b)=>a.dispense<b.dispense?1:-1);
  return {idx,nrec,nid:Object.keys(idx).length};
}

// ---- 每日領藥紀錄檔（S 檔，Big5 制表符分隔）----
// 將「115/06/18」轉成「2026-06-18」，與申報檔解析後的日期格式一致。
function rocSlashToAd(s){
  s=(s||"").trim();
  const m=s.match(/^(\d+)\/(\d{1,2})\/(\d{1,2})$/);
  if(!m) return s;
  return (parseInt(m[1],10)+1911)+"-"+m[2].padStart(2,"0")+"-"+m[3].padStart(2,"0");
}
// 處方箋類別標示：
//  慢性處方箋 → 欄71「序號: N」有值，顯示「慢性處方次數：IC 0N / 次數(欄70)」。
//  一般處方箋 → 序號/次數空白，顯示「一般處方：IC 就診序號(欄24，依就診次數異動)」。
function rxLabelFrom(f){
  const cnt=(f[70]||"").replace(/[^0-9]/g,"");   // 次數
  const seq=(f[71]||"").replace(/[^0-9]/g,"");   // 序號
  if(seq) return "慢性處方次數：IC "+seq.padStart(2,"0")+(cnt?" / "+cnt:"");
  const code=((f[24]||"").trim()||(f[61]||"").trim()).replace(/^IC/i,"").trim();
  return "一般處方：IC "+(code||"—");
}
// 解析 S 檔；以「處方調劑\t」為每筆紀錄起點（自由文字欄位內含換行，不能單純以行切割）。
// 同一張處方箋的多項藥品以流水號(欄2)聚合為一次領藥；領藥時:分取欄92/93/94中
// 最晚的時間戳（處方建立→調劑→領藥，最後一關即實際交付時間）。
function parseTimeText(text){
  const blocks=text.split(/\r?\n(?=處方調劑\t)/);
  const groups={};
  for(const raw of blocks){
    if(!raw || raw.indexOf("\t")<0) continue;
    const f=raw.split("\t");
    if(f.length<95) continue;
    const id=(f[6]||"").trim().toUpperCase();
    const code=(f[35]||"").trim().toUpperCase();
    if(!id || !code) continue;
    const dispKey=(f[4]||"").trim().replace(/\//g,"");   // 1150618
    // 領藥(交付)時間：取欄92/93/94中最晚的時間戳（處方建立→調劑→領藥的最後一關）。
    // 慢箋提前領藥時，申報日(欄4)可能晚於實際交付日，故不限定日期、取最晚者。
    let best="";
    for(const ci of [92,93,94]){
      const v=(f[ci]||"").trim();
      if(/^\d{13}$/.test(v) && v>best) best=v;
    }
    const hm=best?best.slice(7,9)+":"+best.slice(9,11):"";
    const pdate=best?rocSlashToAd(best.slice(0,3)+"/"+best.slice(3,5)+"/"+best.slice(5,7)):"";
    const uid=id+"|"+dispKey+"|"+(f[2]||"").trim();
    let g=groups[uid];
    if(!g){
      g=groups[uid]={uid:uid,id:id,name:(f[7]||"").trim(),
        birth:rocSlashToAd(f[59]),visit:rocSlashToAd(f[5]),
        dispense:rocSlashToAd(f[4]),dept:(f[25]||"").trim(),
        hosp:(f[3]||"").trim(),hospname:(f[65]||"").trim(),
        points:(f[13]||"").trim(),time:hm,pdate:pdate,bestTs:best,
        rxType:rxLabelFrom(f),src:"S",drugs:[]};
    } else if(best>g.bestTs){ g.bestTs=best; g.time=hm; g.pdate=pdate; }
    g.drugs.push({kind:"1",code:code,
      name:(f[81]||"").trim()||DRUGS[code]||(f[36]||"").trim(),
      dose:(f[37]||"").trim(),freq:(f[39]||"").trim(),
      route:(f[40]||"").trim(),days:(f[38]||"").trim(),qty:(f[41]||"").trim()});
  }
  const idx={}; let np=0;
  for(const uid in groups){ const g=groups[uid]; (idx[g.id]=idx[g.id]||[]).push(g); np++; }
  return {idx:idx,np:np};
}
// 把一次解析結果併入全域索引，以 uid 去重（同筆重覆匯入不會重算）。
function addTimeRecs(parsed){
  let added=0;
  for(const id in parsed.idx){
    for(const rec of parsed.idx[id]){
      if(STIME_SEEN[rec.uid]) continue;
      STIME_SEEN[rec.uid]=1;
      (STIMES_BY_ID[id]=STIMES_BY_ID[id]||[]).push(rec);
      added++;
    }
  }
  return added;
}
function hasTimes(){ return Object.keys(STIMES_BY_ID).length>0; }
// 合併某身分證的申報檔紀錄與每日領藥紀錄：
//  日期+院所相同者，把時:分補到申報檔該筆；否則把每日紀錄當作獨立一次領藥列入。
function combinedRecs(k){
  const base=(INDEX&&INDEX[k])?INDEX[k]:[];
  const out=base.map(r=>Object.assign({},r));
  const keyOf=r=>(r.id||"")+"|"+r.dispense+"|"+r.hosp;
  const map={}; out.forEach(r=>{map[keyOf(r)]=r;});
  for(const sr of (STIMES_BY_ID[k]||[])){
    const m=map[keyOf(sr)];
    if(m){ if(sr.time){ m.time=sr.time; m.pdate=sr.pdate; } if(!m.hospname&&sr.hospname) m.hospname=sr.hospname; }
    else out.push(sr);
  }
  out.sort((a,b)=>a.dispense<b.dispense?1:(a.dispense>b.dispense?-1:0));
  return out;
}
function updateBar(){
  const on=(INDEX&&Object.keys(INDEX).length)||hasTimes();
  document.getElementById("bar").classList.toggle("on",!!on);
}

async function unzipFirstXml(buf){
  // 解析 ZIP（以中央目錄為準），取第一個檔，必要時用 DecompressionStream 解壓 deflate
  const dv=new DataView(buf), u8=new Uint8Array(buf);
  let eocd=-1;
  for(let k=u8.length-22;k>=0;k--){ if(dv.getUint32(k,true)===0x06054b50){eocd=k;break;} }
  if(eocd<0) throw new Error("不是有效的 ZIP 檔");
  const cd=dv.getUint32(eocd+16,true);
  const method=dv.getUint16(cd+10,true);
  const csize=dv.getUint32(cd+20,true);
  const localOff=dv.getUint32(cd+42,true);
  const lnameLen=dv.getUint16(localOff+26,true);
  const lextraLen=dv.getUint16(localOff+28,true);
  const dataStart=localOff+30+lnameLen+lextraLen;
  const comp=u8.slice(dataStart,dataStart+csize);
  if(method===0) return comp.buffer.slice(comp.byteOffset,comp.byteOffset+comp.byteLength);
  if(method===8){
    if(typeof DecompressionStream==="undefined")
      throw new Error("瀏覽器不支援解壓，請改用新版 Chrome/Edge/Safari，或直接匯入未壓縮的 DRUGT.xml");
    const ds=new DecompressionStream("deflate-raw");
    const stream=new Blob([comp]).stream().pipeThrough(ds);
    return await new Response(stream).arrayBuffer();
  }
  throw new Error("不支援的壓縮方式 method="+method);
}

document.getElementById("file").addEventListener("change",async function(e){
  const f=e.target.files[0]; if(!f) return;
  const st=document.getElementById("st");
  st.className="status"; st.textContent="解析中… "+f.name;
  try{
    let buf=await f.arrayBuffer();
    const sig=new DataView(buf).getUint32(0,true);
    const isZip = /\.zip$/i.test(f.name) || sig===0x04034b50;
    if(isZip){ st.textContent="解壓中… "+f.name; buf=await unzipFirstXml(buf); }
    let text=new TextDecoder("big5").decode(new Uint8Array(buf));
    if(!/<ddata>/.test(text)) text=new TextDecoder("utf-8").decode(new Uint8Array(buf));
    if(!/<ddata>/.test(text)) throw new Error("檔案內找不到 <ddata> 資料，請確認是健保調劑申報檔");
    const r=parseXML(text);
    INDEX=r.idx;
    st.className="status ok";
    st.textContent="✅ 已載入 "+f.name+"　"+r.nrec+" 筆紀錄 / "+r.nid+" 位";
    document.getElementById("bar").classList.add("on");
    document.getElementById("clrBtn").style.display="inline-block";
    document.getElementById("q").focus();
    document.getElementById("out").innerHTML="";
    try{ await idbPut("last",{name:f.name,nrec:r.nrec,nid:r.nid,idx:r.idx,ts:Date.now()}); }catch(e){}
  }catch(err){
    st.className="status warn"; st.textContent="⚠️ 解析失敗："+err.message;
  }
});

// 匯入一批每日領藥紀錄檔（資料夾或多選共用）：只挑 .txt/.zip，逐檔解析、壞檔略過。
async function importTimeFiles(files){
  const tst=document.getElementById("tst");
  files=files.filter(f=>/\.(txt|zip)$/i.test(f.name));
  if(!files.length){ tst.className="status warn"; tst.textContent="⚠️ 選取範圍內沒有 .txt／.zip 檔可匯入"; return; }
  let totalAdded=0, okFiles=0, skipped=0;
  for(let i=0;i<files.length;i++){
    const f=files[i];
    tst.className="status"; tst.textContent="解析中（"+(i+1)+"/"+files.length+"）… "+f.name;
    try{
      let buf=await f.arrayBuffer();
      const sig=new DataView(buf).getUint32(0,true);
      if(/\.zip$/i.test(f.name)||sig===0x04034b50){ buf=await unzipFirstXml(buf); }
      let text=new TextDecoder("big5").decode(new Uint8Array(buf));
      if(text.indexOf("處方調劑")<0) text=new TextDecoder("utf-8").decode(new Uint8Array(buf));
      if(text.indexOf("處方調劑")<0) throw new Error("非領藥紀錄檔");
      totalAdded+=addTimeRecs(parseTimeText(text)); okFiles++;
    }catch(err){ skipped++; }
  }
  if(!okFiles){ tst.className="status warn"; tst.textContent="⚠️ 選取的 "+files.length+" 個檔都無法解析為領藥紀錄檔"; return; }
  let nrec=0; for(const id in STIMES_BY_ID) nrec+=STIMES_BY_ID[id].length;
  const nid=Object.keys(STIMES_BY_ID).length;
  try{ await idbPut("stimes",{byId:STIMES_BY_ID,seen:STIME_SEEN,ts:Date.now()}); }catch(e){}
  tst.className="status ok";
  tst.textContent="✅ 已匯入 "+okFiles+" 個檔"+(skipped?"（略過 "+skipped+" 個）":"")+
    "　本次 +"+totalAdded+" 筆，累計 "+nrec+" 次領藥 / "+nid+" 位";
  document.getElementById("tclrBtn").style.display="inline-block";
  updateBar();
  if(document.getElementById("q").value.trim()) run();
}
document.getElementById("tfile").addEventListener("change",function(e){ importTimeFiles([...e.target.files]); e.target.value=""; });
document.getElementById("tdir").addEventListener("change",function(e){ importTimeFiles([...e.target.files]); e.target.value=""; });

async function clearTimes(){
  try{ await idbDel("stimes"); }catch(e){}
  STIMES_BY_ID={}; STIME_SEEN={};
  document.getElementById("tclrBtn").style.display="none";
  const tst=document.getElementById("tst"); tst.className="status";
  tst.textContent="已清除領藥時間檔　•　含每筆精確領藥時:分，可累積多日";
  updateBar();
  if(document.getElementById("q").value.trim()) run();
}

function run(){
  const k=document.getElementById("q").value.trim().toUpperCase();
  const out=document.getElementById("out");
  if(!INDEX && !hasTimes()){ return; }
  if(!k){ out.innerHTML='<div class="empty">請輸入身分證字號</div>'; return; }
  const recs=combinedRecs(k);
  if(!recs.length){ out.innerHTML='<div class="empty">查無此身分證字號：'+esc(k)+'</div>'; return; }
  let h='<p class="person">'+esc(recs[0].name)+' <b>'+esc(k)+'</b>　出生 '+esc(recs[0].birth)+'　共 '+recs.length+' 次領藥</p>';
  for(const r of recs){
    h+='<div class="card"><div class="date">'+esc(r.dispense)+'　領藥'+(r.time?' <span class="time">'+esc(r.time)+'</span>':'')+((r.time&&r.pdate&&r.pdate!==r.dispense)?' <span class="tag">實際領藥 '+esc(r.pdate)+'</span>':'')+(r.src==='S'?' <span class="tag">每日明細</span>':'')+'</div>';
    h+='<div class="meta">就醫日 '+esc(r.visit)+'　•　科別 '+esc(DEPT[r.dept]||r.dept||'—')+
       '　•　院所 '+esc(r.hospname||r.hosp)+(r.hospname?' ('+esc(r.hosp)+')':'')+
       '　•　申報點數 '+esc(r.points)+'</div>';
    if(r.rxType) h+='<div class="meta rx">'+esc(r.rxType)+'</div>';
    const meds=r.drugs.filter(d=>d.kind==="1");
    if(meds.length){
      h+='<table><tr><th>藥品代碼</th><th>藥名</th><th>用量</th><th>用法</th><th>途徑</th><th>日數</th><th>總量</th></tr>';
      for(const d of meds){
        h+='<tr><td><span class="tag">'+esc(d.code)+'</span></td><td class="dname">'+
           (d.name?esc(d.name):'<span style="color:var(--muted)">—</span>')+
           '</td><td>'+esc(d.dose)+'</td><td>'+esc(d.freq)+'</td><td>'+esc(d.route)+
           '</td><td>'+esc(d.days)+'</td><td>'+esc(d.qty)+'</td></tr>';
      }
      h+='</table>';
    }
    h+='</div>';
  }
  out.innerHTML=h;
}
// ---- 來源持久化（IndexedDB，僅存於本機瀏覽器）----
function idbOpen(){return new Promise((res,rej)=>{const r=indexedDB.open("yaoju_db",1);
  r.onupgradeneeded=()=>r.result.createObjectStore("src");
  r.onsuccess=()=>res(r.result); r.onerror=()=>rej(r.error);});}
async function idbPut(k,v){const db=await idbOpen();return new Promise((res,rej)=>{
  const tx=db.transaction("src","readwrite"); tx.objectStore("src").put(v,k);
  tx.oncomplete=()=>res(); tx.onerror=()=>rej(tx.error);});}
async function idbGet(k){const db=await idbOpen();return new Promise((res,rej)=>{
  const tx=db.transaction("src","readonly"); const rq=tx.objectStore("src").get(k);
  rq.onsuccess=()=>res(rq.result); rq.onerror=()=>rej(rq.error);});}
async function idbDel(k){const db=await idbOpen();return new Promise((res,rej)=>{
  const tx=db.transaction("src","readwrite"); tx.objectStore("src").delete(k);
  tx.oncomplete=()=>res(); tx.onerror=()=>rej(tx.error);});}

async function clearSrc(){
  try{ await idbDel("last"); }catch(e){}
  INDEX=null;
  updateBar();
  document.getElementById("clrBtn").style.display="none";
  const st=document.getElementById("st"); st.className="status";
  st.textContent="已清除預設　•　請選擇健保調劑申報檔 DRUGT.xml 或 DRUGT.zip";
  document.getElementById("out").innerHTML="";
  document.getElementById("q").value="";
}

window.addEventListener("DOMContentLoaded",async()=>{
  try{
    const saved=await idbGet("last");
    if(saved && saved.idx){
      INDEX=saved.idx;
      const st=document.getElementById("st"); st.className="status ok";
      st.textContent="✅ 已預設上次來源 "+saved.name+"　"+saved.nrec+" 筆 / "+saved.nid+" 位（選新檔可更換）";
      document.getElementById("clrBtn").style.display="inline-block";
    }
  }catch(e){}
  try{
    const savedT=await idbGet("stimes");
    if(savedT && savedT.byId){
      STIMES_BY_ID=savedT.byId; STIME_SEEN=savedT.seen||{};
      let nrec=0; for(const id in STIMES_BY_ID) nrec+=STIMES_BY_ID[id].length;
      const tst=document.getElementById("tst"); tst.className="status ok";
      tst.textContent="✅ 已預設領藥時間檔　累計 "+nrec+" 次領藥 / "+Object.keys(STIMES_BY_ID).length+" 位（可再匯入累積）";
      document.getElementById("tclrBtn").style.display="inline-block";
    }
  }catch(e){}
  updateBar();
});

document.getElementById("q").addEventListener("keydown",e=>{ if(e.key==="Enter") run(); });
</script>
</body>
</html>"""


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
